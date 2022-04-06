import airflow
from airflow.utils.dates import days_ago
from airflow import DAG
from airflow.operators.docker_operator import DockerOperator
from airflow.operators.python_operator import PythonOperator
from airflow.operators.bash_operator import BashOperator
import requests
import pandas as pd
from datetime import timedelta
from openpyxl import load_workbook
from docker.types import Mount
import os
from sys import api_version

 
default_args = {
    'owner': 'Victor',    
    'start_date': airflow.utils.dates.days_ago(2),
    'retry_delay': timedelta(seconds=10),
    'retries': 1,
    }

# Primeiramente é necessário importar o arquivo da fonte.

def extractFile():
    dls = "https://github.com/raizen-analytics/data-engineering-test/raw/master/assets/vendas-combustiveis-m3.xls"
    resp = requests.get(dls, allow_redirects=True)

    output = open('/diretorio/desafio.xls', 'wb')
    output.write(resp.content)
    #output.close()

# Com o arquivo em mãos, foi necessário transformalo para o formato XLSX para assim poder acessar os dados da tabela pivot;
# Para isso foi utilizado um DockerOperator que extrai o arquivo XLS para o libreoffice, onde atráves de um comando Linux, o transforma para o formato desejado;
# Com o arquivo transformado, foi utilizado a bibíoteca openpyxl para extrair o conteudo interno, e assim começar o manipulatingamento.

def extractSheet(sheet_name,file_name,destiny):
    workBook = load_workbook(destiny)
    sheets = workBook.sheetnames
    for sheet in sheets:
        if sheet != sheet_name:
            workBook.remove(workBook.get_sheet_by_name(sheet))
    
    workBook[sheet_name].sheet_state = 'visible'
    workBook.save(file_name)

def extract_diesel_uf_and_type():
    extractSheet('DPCache_m3','extract_diesel_by_uf_and_type.xlsx','/diretorio/desafio_1.xlsx')    

def extract_oil_derivative_fuels_by_uf_and_product():
    extractSheet('DPCache_m3 2','extract_oil_derivative_fuels_by_uf_and_product.xlsx','/diretorio/desafio_1.xlsx')

# Traduzindo "Estado" para "UF".

def manipulatingData(file_name,end_name):
    combustivel = pd.read_excel(file_name)
    combustivel = combustivel.astype({'ANO': 'str'})
    combustivel['ESTADO'] = combustivel['ESTADO'].map({'ACRE':'AC',
    'AlAGOAS': 'AL',
    'AMAPÁ':'AP',
    'AMAZONAS':'AM',
    'BAHIA':'BA',
    'CEARÁ':'CE',
    'DISTRITO FEDERAL':'DF' ,
    'ESPÍRITO SANTO':'ES',
    'GOIÁS':'GO',
    'MARANHÃO':'MA',
    'MATO GROSSO':'MT' ,
    'MATO GROSSO DO SUL': 'MS',
    'MINAS GERAIS':'MG',
    'PARÁ':'PA',
    'PARAÍBA':'PB',
    'PARANÁ':'PR',
    'PERNAMBUCO':'PE',
    'PIAUÍ':'PI',
    'RIO DE JANEIRO':'RJ',
    'RIO GRANDE DO NORTE':'RN',
    'RIO GRANDE DO SUL':'RS',
    'RONDÔNIA':'RO',
    'RORAIMA':'RR',
    'SANTA CATARINA':'SC',
    'SÃO PAULO':'SP',
    'SERGIPE':'SE',
    'TOCANTINS':'TO'})
    
# Agrupando a coluna "MES" para formar a coluna "Volume".
# Agrupando os df para formatar a nova coluna "Ano_Mes".

    mes1 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Jan']].rename(columns={'Jan': 'volume'})
    mes1['ANO_MES'] = mes1['ANO'] + '-' + '01'
    mes2 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Fev']].rename(columns={'Fev': 'volume'})
    mes2['ANO_MES'] = mes2['ANO'] + '-' + '02'
    mes3 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Mar']].rename(columns={'Mar': 'volume'})
    mes3['ANO_MES'] = mes3['ANO'] + '-' + '03'
    mes4 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Abr']].rename(columns={'Abr': 'volume'})
    mes4['ANO_MES'] = mes4['ANO'] + '-' + '04'
    mes5 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Mai']].rename(columns={'Mai': 'volume'})
    mes5['ANO_MES'] = mes5['ANO'] + '-' + '05'
    mes6 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Jun']].rename(columns={'Jun': 'volume'})
    mes6['ANO_MES'] = mes6['ANO'] + '-' + '06'
    mes7 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Jul']].rename(columns={'Jul': 'volume'})
    mes7['ANO_MES'] = mes7['ANO'] + '-' + '07'
    mes8 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Ago']].rename(columns={'Ago': 'volume'})
    mes8['ANO_MES'] = mes8['ANO'] + '-' + '08'
    mes9 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Set']].rename(columns={'Set': 'volume'})
    mes9['ANO_MES'] = mes9['ANO'] + '-' + '09'
    mes10 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Out']].rename(columns={'Out': 'volume'})
    mes10['ANO_MES'] = mes10['ANO'] + '-' + '10'
    mes11 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Nov']].rename(columns={'Nov': 'volume'})
    mes11['ANO_MES'] = mes11['ANO'] + '-' + '11'
    mes12 = combustivel[['ANO','ESTADO','COMBUSTÍVEL','UNIDADE','Dez']].rename(columns={'Dez': 'volume'})
    mes12['ANO_MES'] = mes12['ANO'] + '-' + '12'
    
    volume = pd.concat(([mes1, mes2, mes3, mes4, mes5, mes6, mes7, mes8, mes9, mes10, mes11, mes12]))
    volume.drop(columns=['ANO'])
    volume = volume.rename(columns={'ANO_MES':'year_month','COMBUSTÍVEL': 'product','ESTADO': 'uf','UNIDADE': 'unit'})
    volume['created_at'] = pd.to_datetime('today')
    volume.astype({'year_month' : 'datetime64','uf' : 'string','volume' : 'float', 'product' : 'string','unit' : 'string'})
    volume.to_parquet(end_name, compression='snappy', partition_cols=['uf','year_month'])


def manipulating_diesel_by_uf_and_type():
    manipulatingData('extract_diesel_by_uf_and_type.xlsx','extract_diesel_by_uf_and_type.parquet')

def manipulating_oil_derivative_fuels_by_uf_and_product():
    manipulatingData('extract_oil_derivative_fuels_by_uf_and_product.xlsx','extract_oil_derivative_fuels_by_uf_and_product.parquet')


with DAG(
    dag_id = 'Desafio_raizen_1',
    default_args=default_args,
    start_date=days_ago(2),
    dagrun_timeout=timedelta(minutes=60),
    max_active_runs=1,
) as dag:

    extractFiles = PythonOperator(
        task_id="extractFile",
        python_callable= extractFile
    )


    extract_diesel_uf_types = PythonOperator(
        task_id="extract_diesel_uf_and_type",
        python_callable = extract_diesel_uf_and_type
    )

    extract_oil_derivative_fuels_by_uf_products = PythonOperator(
        task_id="extract_oil_derivative_fuels_by_uf_and_product",
        python_callable = extract_oil_derivative_fuels_by_uf_and_product
    )

    manipulating_diesel_by_uf_and_types = PythonOperator(
        task_id="manipulating_diesel_by_uf_and_type",
        python_callable=manipulating_diesel_by_uf_and_type
    )
    manipulating_oil_derivative_fuels_by_uf_and_products = PythonOperator(
        task_id="manipulating_oil_derivative_fuels_by_uf_and_product",
        python_callable=manipulating_oil_derivative_fuels_by_uf_and_product
    )
        
    extractFiles >> extract_oil_derivative_fuels_by_uf_products
    extractFiles >> extract_diesel_uf_types
    extract_diesel_uf_types >> manipulating_diesel_by_uf_and_types
    extract_oil_derivative_fuels_by_uf_products >> manipulating_oil_derivative_fuels_by_uf_and_products


    






